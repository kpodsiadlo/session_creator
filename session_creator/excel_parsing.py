import string
from openpyxl import load_workbook
from user_input import UserInput


def get_filenames_from_excel_column(user_input):
    """Return a list of entries in a given column"""

    column_letter_lowercase = user_input.column.lower()
    column_number = int(ord(column_letter_lowercase) - 96)
    # get the file
    wb = load_workbook(user_input.list_file_path)
    # get the sheet
    ws = wb['Sheet1']
    # get the values
    file_list = [ws.cell(row=i, column=column_number).value
                 for i in range(user_input.start_row, user_input.last_row+1)]

    return file_list
